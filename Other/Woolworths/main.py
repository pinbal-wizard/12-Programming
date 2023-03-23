from bs4 import BeautifulSoup

from datetime import date
import selenium
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
from openpyxl import load_workbook


def scrapePage(driver):
    products = {}
    productTiles = driver.find_elements(By.CLASS_NAME,"product-tile-v2")
    for product in productTiles:
        title = product.find_element(By.CLASS_NAME,"product-title-link").text
        try:
            price = product.find_element(By.CLASS_NAME,"product-tile-price")
            pricenumber = price.text.split('\n')[0].replace("$","")
        except:
            pricenumber = -1
        products[title] = pricenumber
        print(f"{title}: {pricenumber}")
    return products


def nextPage(driver):
    button = driver.find_element(By.CLASS_NAME,"paging-next")
    button.click()

    time.sleep(10)
    driver.execute_script("""
    var element = document.querySelector(".container-carousel");
    if (element)
        element.parentNode.removeChild(element);
    """)


def saveData(products, page):
    print("\n\nsaving...")
    row = (today - start_date).days + 2
    workbook = load_workbook("data.xlsx")
    sheet = workbook[page]
    
    sheet.cell(row = row, column = 1).value = today

    productCol = 0
    for i in range(len(products)):
        for cell in sheet.iter_cols(min_col=2, max_col=len(products) + 1, max_row=1, min_row=1):
            if cell[0].value == None:
                cell[0].value = list(products.keys())[i]

            if cell[0].value == list(products.keys())[i]:
                productCol = cell[0].column 
                sheet.cell(row = row,column = productCol).value = list(products.values())[i]
                print(f"product {i} name @ {list(products.keys())[i]} col:{cell[0].column} row:{cell[0].row}")
                print(f"product {i} price @ {list(products.values())[i]} col:{productCol} row:{row}")
                break
    workbook.save(r'data.xlsx')
    print("Finished Saving")


def scrapeAllPage(URL):
    driver = webdriver.Firefox()
    driver.get(URL)

    time.sleep(5)
    driver.execute_script("""
    var element = document.querySelector(".container-carousel");
    if (element)
        element.parentNode.removeChild(element);
    """)
    time.sleep(5)
    for i in range(100): 
        try:
            products.update(scrapePage(driver))
        except Exception as e:
            print(str(e.__class__))
            driver.refresh()
            print("Element not found refreshing... \n\n\n")
            time.sleep(10)
            try:
                products.update(scrapePage(driver))
            except:
                break
        try:
            nextPage(driver)
        except selenium.common.exceptions.NoSuchElementException:
            print("Error clicking next page")
            driver.refresh()
            time.sleep(10)
            try:
                nextPage(driver)
            except:
                print("two times filed assuming end of pages")
                break
    saveData(products,URL.split("/")[-1])
    


URLS = []

"""
    already done
        "https://www.woolworths.com.au/shop/browse/fruit-veg",
        "https://www.woolworths.com.au/shop/browse/meat-seafood-deli",
        "https://www.woolworths.com.au/shop/browse/bakery",
        "https://www.woolworths.com.au/shop/browse/dairy-eggs-fridge",
        "https://www.woolworths.com.au/shop/browse/snacks-confectionery",
        "https://www.woolworths.com.au/shop/browse/lunch-box",
        "https://www.woolworths.com.au/shop/browse/freezer",
        "https://www.woolworths.com.au/shop/browse/drinks",

    fucking massive
        "https://www.woolworths.com.au/shop/browse/pantry",

    who cares    
        "https://www.woolworths.com.au/shop/browse/health-wellness",
        "https://www.woolworths.com.au/shop/browse/liquor",
        "https://www.woolworths.com.au/shop/browse/pet",
        "https://www.woolworths.com.au/shop/browse/baby",
        "https://www.woolworths.com.au/shop/browse/beauty-personal-care",
        "https://www.woolworths.com.au/shop/browse/household"

"""

wb = load_workbook("data.xlsx")
for i in URLS:
    wb.create_sheet(i.split("/")[-1])
wb.save("data.xlsx")
start_date = date.fromisocalendar(2023, 12, 4)
today = date.today()
print((today - start_date).days + 2)

for pageName in URLS:
    products = {}
    scrapeAllPage(pageName)

